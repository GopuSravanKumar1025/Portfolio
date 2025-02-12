from datetime import datetime, timedelta
import requests
import json
import pandas as pd
import os
from pathlib import Path
import feedparser
from urllib.parse import quote
import urllib.parse
from openpyxl import load_workbook
from docx import Document
import re

class ClinicalTrial:

    @staticmethod #DateTime
    def datetime_():
        Today = datetime.now()
        TwoMonthsAgo = Today - timedelta(days=60)
        return Today, TwoMonthsAgo
    @staticmethod #5th Integrated 2nd Part
    def ChangePlaceholders(NameOfEmail, Condition, Body, NCTID, Subject):
        Template = re.sub(r'\[Contact Name\]', NameOfEmail, Body)
        Template = re.sub(r'\[Condition\]', Condition, Template)
        SubjectTemplate = re.sub(r'\[Contact Name\]', NameOfEmail, Subject)
        SubjectTemplate = re.sub(r'\[NCT Number\]', NCTID, SubjectTemplate)
        return Template, SubjectTemplate
    @staticmethod #5th Integrated
    def Req_Meth_Exrt_Data(_Company_Name_, headers, Subject, Body):
        AppendBody = []
        encoded_company_name = quote(_Company_Name_)
        RequestedURL = f'https://clinicaltrials.gov/api/v2/studies?format=json&query.spons={encoded_company_name}' #Api url
        response = requests.get(RequestedURL, headers=headers)
        if response.status_code == 200:
            ClinicalData = response.json()
            ScrappedStudies = ClinicalData.get('studies', [])
            if ScrappedStudies:
                for data in ScrappedStudies:
                    identification_module = data['protocolSection']['identificationModule']
                    Email_Module = ""
                    NameOfEmail = ""
                    BodyOfMail = ""
                    contacts_locations_module = data.get('protocolSection', {}).get('contactsLocationsModule', {})
                    central_contacts = contacts_locations_module.get('centralContacts', [])
                    OverAllStatus = contacts_locations_module.get('overallOfficials', [])
                    if central_contacts and isinstance(central_contacts, list):
                        Email_Module = central_contacts[0].get('email', '')
                        NameOfEmail = central_contacts[0].get('name', '')
                    elif OverAllStatus and isinstance(OverAllStatus, list):
                        Email_Module = OverAllStatus[0].get('email', '')
                        NameOfEmail = OverAllStatus[0].get('name', '')
                    conditions_module = data.get('protocolSection', {}).get('conditionsModule', {})
                    Condition = conditions_module.get('conditions', [])[0] if conditions_module.get('conditions') else ''
                    official_title = identification_module.get('officialTitle')
                    NCTID = data['protocolSection']['identificationModule']['nctId']
                    body = {
                                    'Organization': data['protocolSection']['identificationModule']['organization']['fullName'],
                                    'NCTID': NCTID,
                                    'BriefTitle': data['protocolSection']['identificationModule']['briefTitle'],
                                    'OfficialTitle' : official_title,
                                    'Status': data['protocolSection']['statusModule']['overallStatus'],
                                    'Link': f'https://clinicaltrials.gov/study/{data["protocolSection"]["identificationModule"]["nctId"]}', #general url
                                    'Conditions': Condition
                                }
                    if '@' in Email_Module:
                        BodyOfMail, subject = ClinicalTrial.ChangePlaceholders(NameOfEmail, Condition, Body, NCTID, Subject)
                        BCCMail = 'john.ahmed@mycardium.com'
                        subject_encoded = urllib.parse.quote(subject)
                        body_encoded = urllib.parse.quote(BodyOfMail)
                        bcc_encoded = urllib.parse.quote(BCCMail)
                        mailto_link = f"mailto:{Email_Module}?subject={subject_encoded}&body={body_encoded}&Bcc={BCCMail}"
                        body.update({"Email" : f"{mailto_link}"})
                        AppendBody.append(body)
                    else:
                        body.update({"Email" : ""})
                        AppendBody.append(body)
        return AppendBody
    @staticmethod #5th
    def ExtractData(GetHighlightedData, headers, search_areas, Subject, Body):
        Customers_Data = []
        Prospects_Data = []
        for item in GetHighlightedData:
            if 'highlightedData' in item:
                Extract_highlighted_company_name = item['highlightedData']
                Get_Data = ClinicalTrial.Req_Meth_Exrt_Data(Extract_highlighted_company_name, headers, Subject, Body)
                for data in Get_Data:
                    for i in search_areas:
                        if i in data['Conditions']:
                            Customers_Data.extend(Get_Data)
            else:
                if 'Normal_Data' in item:
                    Extract_highlighted_company_name = item['Normal_Data']
                    Get_Data_Prospects = ClinicalTrial.Req_Meth_Exrt_Data(Extract_highlighted_company_name, headers, Subject, Body)
                    for data in Get_Data_Prospects:
                        for i in search_areas:
                            if i in data['Conditions']:
                                Prospects_Data.extend(Get_Data_Prospects)
                    
        return Customers_Data, Prospects_Data
    @staticmethod #4th
    def wordocument(WordDoc):
        subject = ""
        body = ""
        for i, para in enumerate(WordDoc.paragraphs):
            if i == 0:
                subject = para.text
            else:
                body += para.text + "\n\n"
        return subject, body
    @staticmethod #3rd integrated
    def FilterNewsLinks(company_name, search_areas):
        Today, Since_TwoMonths = ClinicalTrial.datetime_()
        AppendLinks = []
        if company_name:
            encoded_company_name = quote(company_name)
            APIURl = f'https://news.google.com/rss/search?q={encoded_company_name}+after:{Since_TwoMonths.date()}+before:{Today.date()}'
            feed = feedparser.parse(APIURl)
            if feed.status == 302:
                for entry in feed.entries:
                    for area in search_areas:
                        if area.lower() in entry.title.lower():
                            body = {
                                'CompanyName' : company_name,
                                "title": entry.title,
                                "link": entry.link,
                                "NewsAgent" : entry.source.title if 'source' in entry else 'Unknown',
                                "published": entry.published
                            }
                            AppendLinks.append(body)
            else:
                return 'Status Invalid'
        return AppendLinks
    
    @staticmethod #3rd
    def NewsArticles(GetHighlightedData, search_areas):
        Customers_Links = []
        Prospects_Links = []
        for item in GetHighlightedData:
            if 'highlightedData' in item:
                highlightedData_company_name = item['highlightedData']
                CustomersLinks = ClinicalTrial.FilterNewsLinks(highlightedData_company_name, search_areas)
                Customers_Links.extend(CustomersLinks)
            else:
                if 'Normal_Data' in item:
                    Prospect_company_name = item['Normal_Data']
                    ProspectsLinks = ClinicalTrial.FilterNewsLinks(Prospect_company_name, search_areas)
                    Prospects_Links.extend(ProspectsLinks)
        return Customers_Links, Prospects_Links
                    
    @staticmethod #2nd
    def HighlightedData(GetExcelFileName, sheet_names):
        default_bg_color = '00000000'
        SepratedList = []
        for sheet_name in sheet_names:
            if 'Tab' in sheet_name:
                df = pd.read_excel(GetExcelFileName, sheet_name=sheet_name)
                wb = load_workbook(GetExcelFileName)
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
                    for cell in row:
                        value = cell.value
                        bgColor = cell.fill.start_color.rgb if cell.fill.start_color and cell.fill.start_color.rgb else default_bg_color
                        if bgColor != default_bg_color:
                            SepratedList.append({
                                'highlightedData':  value
                            })
                        else:
                             SepratedList.append({
                                'Normal_Data':  value
                            })
        return SepratedList
    @staticmethod #1st
    def FindFile(Baseurl, fileName):
        GetPath = Path(Baseurl)
        for file in GetPath.rglob(fileName):
            if file:
                return file
            else:
                return None
    @staticmethod
    def fetchData():
        username = os.getlogin()
        Baseurl = rf'C:\Users\{username}'
        fileName_Excel = 'ProspectList.xlsx'
        fileName_Json = 'SearchData.json'
        WordDocument = 'Prospect Email Template v1.1 Dashboard.docx'
        GetExcelFileName = ClinicalTrial.FindFile(Baseurl, fileName_Excel)
        wb = load_workbook(GetExcelFileName)
        sheet_names = wb.sheetnames
        GetJsonFileName = ClinicalTrial.FindFile(Baseurl, fileName_Json)
        with open(GetJsonFileName, 'r') as file:
            JSONdata = json.load(file)
        search_areas = JSONdata.get('SearchArea', [])
        WordDoc = Document(ClinicalTrial.FindFile(Baseurl, WordDocument))
        Subject, Body = ClinicalTrial.wordocument(WordDoc)
        GetHighlightedData = ClinicalTrial.HighlightedData(GetExcelFileName, sheet_names)
        wb.close()
        headers = {"accept": "application/json"}
        CustomerLink, ProspectLink = ClinicalTrial.NewsArticles(GetHighlightedData, search_areas)
        CustomersRawData, ProspectsRawData = ClinicalTrial.ExtractData(GetHighlightedData, headers, search_areas, Subject, Body)
        body = {
            "CustomersLink" : CustomerLink,
            "ProspectsLink" : ProspectLink,
            'CustomersRawData' : CustomersRawData,
            'ProspectsRawData' : ProspectsRawData,
        }
        return body